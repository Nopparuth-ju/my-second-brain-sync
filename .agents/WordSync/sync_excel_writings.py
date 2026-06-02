import os
import sys
import re
import openpyxl

# Reconfigure standard output to support UTF-8 (emojis and Thai characters) in Windows console
if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

# Paths Configuration
VAULT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = r"D:\Boss\3) Hobby\3.14) Writing\Writing List.xlsx"
WRITING_OUTPUT_DIR = os.path.join(VAULT_ROOT, "05_OUTPUT", "Writing", "Midgrad")

def parse_desire_tags(desire_str):
    """
    Parse a human desire string (e.g. 'Happiness + Knowledge')
    into a clean Obsidian domain tag string (e.g. '#domain/happiness #domain/knowledge').
    """
    if not desire_str:
        return ""
    # Split by plus, comma, ampersand, or the word 'and'
    parts = re.split(r'[\+,&]|\band\b', str(desire_str))
    tags = []
    for part in parts:
        cleaned = part.strip().lower()
        cleaned = re.sub(r'\s+', '-', cleaned)
        if cleaned:
            tags.append(f"#domain/{cleaned}")
    return " ".join(tags)

def parse_episode_details(name, ep_num):
    """
    Parses series name, normalized EP number, and episode title from name string.
    """
    name = str(name).strip()
    if "|" in name:
        series_part, ep_part = name.split("|", 1)
        series_name = series_part.strip()
        ep_part = ep_part.strip()
        if ":" in ep_part:
            num_part, title_part = ep_part.split(":", 1)
            ep_num_str = re.sub(r'\s+', ' ', num_part.strip())
            episode_title = title_part.strip()
        else:
            # Fallback if no colon exists, e.g. "EP 4 The Hidden Cowardice"
            match = re.match(r'^(EP\s*\d+)(.*)$', ep_part, re.IGNORECASE)
            if match:
                ep_num_str = re.sub(r'\s+', ' ', match.group(1).strip())
                episode_title = match.group(2).strip()
                if episode_title.startswith(":") or episode_title.startswith("-"):
                    episode_title = episode_title[1:].strip()
            else:
                ep_num_str = f"EP {ep_num}"
                episode_title = ep_part
    else:
        series_name = name
        ep_num_str = f"EP {ep_num}"
        episode_title = name
        
    return series_name, ep_num_str, episode_title

def read_sheet_with_propagation(sheet):
    """
    Reads excel sheet, automatically finds header containing 'No' and 'EP',
    and propagates the 'No' value down for empty cells of the same series.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    
    # Locate the header row (contains both 'No' and 'EP' in some casing)
    header_idx = 0
    for idx, row in enumerate(rows):
        row_lower = [str(x).lower().strip() if x is not None else "" for x in row]
        if "no" in row_lower and "ep" in row_lower:
            header_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else f"Col{i}" for i, h in enumerate(rows[header_idx])]
    
    data_rows = []
    current_no = None
    
    for row in rows[header_idx + 1:]:
        # Skip fully empty rows
        if all(x is None for x in row):
            continue
            
        row_dict = {}
        for h, val in zip(headers, row):
            row_dict[h] = val
            
        # Propagate No down
        no_val = row_dict.get('No')
        if no_val is not None and str(no_val).strip() != "":
            try:
                current_no = int(no_val)
            except ValueError:
                current_no = no_val
        row_dict['Propagated_No'] = current_no
        
        # Ensure EP is integer
        ep_val = row_dict.get('EP')
        if ep_val is not None:
            try:
                row_dict['Clean_EP'] = int(ep_val)
            except ValueError:
                row_dict['Clean_EP'] = ep_val
        else:
            row_dict['Clean_EP'] = None
            
        data_rows.append(row_dict)
        
    return headers, data_rows

def generate_featured_moc_file(vault_root, generated_episodes):
    """
    Generates or updates 04_MOC/Featured Writing MOC.md as a premium hub
    for all major writing series, keeping them separate from daily practice.
    """
    moc_path = os.path.join(vault_root, "04_MOC", "Featured Writing MOC.md")
    
    # Group generated episodes by series
    series_groups = {}
    for ep in generated_episodes:
        s_name = ep['series_name']
        if s_name not in series_groups:
            series_groups[s_name] = []
        series_groups[s_name].append(ep)
        
    # Sort episodes within each series by episode number
    for s_name in series_groups:
        series_groups[s_name].sort(key=lambda x: x['ep_num'])
        
    # Construct the MOC content
    doc_lines = []
    doc_lines.append("#type/moc")
    doc_lines.append("#status/processed")
    doc_lines.append("#domain/writing")
    doc_lines.append("#creator/me")
    doc_lines.append("")
    doc_lines.append("# 📚 Featured Writing MOC (ศูนย์รวมผลงานเขียนหลักและซีรีส์เผยแพร่)")
    doc_lines.append("")
    doc_lines.append("> [!NOTE] Map of Content: แผนผังพอร์ตโฟลิโอผลงานเขียนชิ้นเอก")
    doc_lines.append("> - 🎯 **วัตถุประสงค์:** โน้ตแผ่นนี้ทำหน้าที่เป็น **\"จุดแกนกลาง\" (Hub)** รวบรวมงานเขียนชุดหลักและซีรีส์นิยาย/บทความที่คุณนำไปเผยแพร่หรือแชร์บนช่องทางต่างๆ (เช่น Midgrad) เพื่อให้แยกส่วนกับงานฝึกหัดเขียนรายวัน และทำให้ใน **Graph View** เกิดคลัสเตอร์โยงใยสวยงามของซีรีส์แต่ละตอนแยกออกมาเป็นพอร์ตโฟลิโอดั่งดวงดาวรอบๆ แกนกลางนี้!")
    doc_lines.append("")
    doc_lines.append("---")
    doc_lines.append("")
    doc_lines.append("## 🌟 ซีรีส์ผลงานเขียนทั้งหมด (Featured Writing Portfolio)")
    doc_lines.append("")
    doc_lines.append("นี่คือสารบัญชุดซีรีส์ผลงานชิ้นเอกของคุณที่สร้างขึ้นและเรียบเรียงใน `05_OUTPUT/Writing/` สามารถคลิกเปิดเพื่อเข้าถึงสารบัญย่อยหรืออ่านแต่ละตอนได้ทันที:")
    doc_lines.append("")
    
    # Descriptions map for each series
    series_descriptions = {
        "The Brittle Veneer": "นิยายธรรมะเชิงจิตวิทยาที่ผสานสัจธรรมผ่านมุมมองผู้คนและปีศาจในตัวตน",
        "ใช้ชีวิตง่ายๆ ด้วยการรู้กฎแค่ 2 ข้อ": "บทความพัฒนาตนเอง ปลดปล่อยความคิดผ่านสัจธรรมและการค้นหา Core Value",
        "Walk to Wake": "เรื่องเล่าผ่อนคลายจิตใจเชิงธรรมชาติบำบัด เพื่อตื่นรู้และอยู่กับปัจจุบันขณะ"
    }
    
    # Dynamic icons map
    series_icons = {
        "The Brittle Veneer": "🎭",
        "ใช้ชีวิตง่ายๆ ด้วยการรู้กฎแค่ 2 ข้อ": "💡",
        "Walk to Wake": "🚶‍♂️"
    }
    
    for s_name in sorted(series_groups.keys()):
        desc = series_descriptions.get(s_name, "ชุดผลงานเขียนสร้างสรรค์พิเศษของคุณ")
        icon = series_icons.get(s_name, "✍️")
        
        doc_lines.append(f"### {icon} {s_name}")
        doc_lines.append(f"*{desc}*")
        
        for ep in series_groups[s_name]:
            rel_link = f"05_OUTPUT/Writing/Midgrad/{s_name}/{ep['filename_no_ext']}"
            doc_lines.append(f"* [[{rel_link}|📖 {ep['ep_num_str']} - {ep['episode_title']}]]")
        doc_lines.append("")
        
    doc_lines.append("---")
    doc_lines.append("")
    doc_lines.append("## 🎯 สรุปผลลัพธ์และความคิดสร้างสรรค์ (Portfolio Reflections)")
    doc_lines.append("> [!TIP] บันทึกแนวคิด มุมมอง และเป้าหมายของแต่ละซีรีส์ยาวๆ ได้ตรงนี้ เช่น:")
    doc_lines.append("> - **ทิศทางการทำซีรีส์ถัดไป:** (เช่น อยากทำแนว Sci-Fi จิตวิทยา หรือแนววิชาการสั้นแต่อ่านสนุกเพิ่มขึ้น)")
    doc_lines.append("> - **ช่องทางการแชร์/ฟีดแบ็ก:** (เช่น สถิติคนอ่านจาก Midgrad หรือกระแสการตอบรับที่เป็นประโยชน์เพื่อนำมาเกลาโครงเรื่องถัดไป)")
    doc_lines.append("")
    
    doc_content = "\n".join(doc_lines)
    
    # Write MOC file in UTF-8 with BOM format!
    os.makedirs(os.path.dirname(moc_path), exist_ok=True)
    with open(moc_path, "w", encoding="utf-8-sig") as f:
        f.write(doc_content)
        
    print(f"   ✅ Successfully created/updated: 04_MOC/{os.path.basename(moc_path)}")

def sync_writings():
    print("🚀 Starting Featured Writing Series Sync...", flush=True)
    
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Error: Excel file not found at: {EXCEL_PATH}")
        sys.exit(1)
        
    print(f"📂 Loading Excel database: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    
    # 1. Read Sheets
    print("📖 Reading Excel sheets...")
    
    # Publish Detail
    if 'Publish Detail' not in wb.sheetnames:
        print("❌ Error: 'Publish Detail' sheet is missing!")
        sys.exit(1)
    _, publish_rows = read_sheet_with_propagation(wb['Publish Detail'])
    
    # Excerpt
    excerpt_rows = []
    if 'Excerpt' in wb.sheetnames:
        _, excerpt_rows = read_sheet_with_propagation(wb['Excerpt'])
    else:
        print("⚠️ Warning: 'Excerpt' sheet is missing.")
        
    # Midgrad
    midgrad_rows = []
    if 'Midgrad' in wb.sheetnames:
        _, midgrad_rows = read_sheet_with_propagation(wb['Midgrad'])
    else:
        print("⚠️ Warning: 'Midgrad' sheet is missing.")
        
    # 2. Build Excerpt and Midgrad mappings
    excerpt_by_key = {}
    for r in excerpt_rows:
        key = (r['Propagated_No'], r['Clean_EP'])
        excerpt_by_key[key] = r.get('Excerpt')
        
    # Propagate series-level metadata in Midgrad (Human Desire, Objective)
    current_no = None
    current_desire = None
    current_objective = None
    
    midgrad_by_key = {}
    for r in midgrad_rows:
        no_val = r['Propagated_No']
        if no_val != current_no:
            current_no = no_val
            current_desire = r.get('Human Desire')
            current_objective = r.get('Objective')
        else:
            if r.get('Human Desire') is None:
                r['Human Desire'] = current_desire
            if r.get('Objective') is None:
                r['Objective'] = current_objective
                
        key = (r['Propagated_No'], r['Clean_EP'])
        midgrad_by_key[key] = {
            'Human Desire': r.get('Human Desire'),
            'Objective': r.get('Objective'),
            'Status': r.get('Status')
        }
        
    # 3. Process and Write each Episode
    os.makedirs(WRITING_OUTPUT_DIR, exist_ok=True)
    generated_episodes = []
    
    print("\n📝 Generating and formatting stories...")
    for r in publish_rows:
        no = r['Propagated_No']
        ep = r['Clean_EP']
        name = r.get('Name')
        detail = r.get('Detail')
        
        if not name or not detail:
            continue
            
        key = (no, ep)
        
        # Merge with other sheets
        excerpt = excerpt_by_key.get(key, "")
        midgrad_info = midgrad_by_key.get(key, {})
        
        human_desire = midgrad_info.get('Human Desire', "")
        objective = midgrad_info.get('Objective', "")
        
        # Clean values (replace None with empty or dash)
        str_excerpt = str(excerpt).strip() if excerpt is not None else ""
        str_desire = str(human_desire).strip() if human_desire is not None else ""
        str_objective = str(objective).strip() if objective is not None else ""
        
        # Parse Episode details
        series_name, ep_num_str, episode_title = parse_episode_details(name, ep)
        
        # Compile tags
        base_tags = "#type/output #creator/me #status/ready #writing/storytelling #writing/midgrad"
        dynamic_tags = parse_desire_tags(str_desire)
        full_tags = f"{base_tags} {dynamic_tags}".strip()
        
        # Determine file and path names
        # Standard filename: EP {Num} - {Title}.md
        # If series has 1 episode and series name equals episode title, it becomes EP 1 - {Series Name}.md
        filename_no_ext = f"{ep_num_str} - {episode_title}"
        
        # Remove any illegal characters from filename
        filename_no_ext = re.sub(r'[\\/:*?"<>|]', '_', filename_no_ext)
        filename = f"{filename_no_ext}.md"
        
        series_dir = os.path.join(WRITING_OUTPUT_DIR, series_name)
        os.makedirs(series_dir, exist_ok=True)
        
        filepath = os.path.join(series_dir, filename)
        
        # Build document content
        doc_lines = []
        doc_lines.append(full_tags)
        doc_lines.append("")
        doc_lines.append("> [!NOTE] ข้อมูลการเขียน")
        doc_lines.append(f"> - 🎯 **วัตถุประสงค์ (Objective):** {str_objective if str_objective else '-'}")
        doc_lines.append(f"> - 💡 **ความต้องการพื้นฐาน (Human Desire):** {str_desire if str_desire else '-'}")
        doc_lines.append(f"> - 📝 **คำโปรย (Excerpt):** {str_excerpt if str_excerpt else '-'}")
        doc_lines.append("")
        doc_lines.append("---")
        doc_lines.append("")
        
        # If the story doesn't already contain heading, prepend the episode title
        doc_lines.append(f"# {ep_num_str} - {episode_title}")
        doc_lines.append("")
        doc_lines.append(str(detail).strip())
        doc_lines.append("")
        
        doc_content = "\n".join(doc_lines)
        
        # Write to file in UTF-8 with BOM format!
        with open(filepath, "w", encoding="utf-8-sig") as f:
            f.write(doc_content)
            
        print(f"   ✅ Saved: 05_OUTPUT/Writing/Midgrad/{series_name}/{filename}")
        
        generated_episodes.append({
            'series_name': series_name,
            'ep_num': ep,
            'ep_num_str': ep_num_str,
            'episode_title': episode_title,
            'filename_no_ext': filename_no_ext
        })
        
    print(f"\n🎉 Successfully processed and saved {len(generated_episodes)} stories!")
    
    # 4. Generate/Update Featured Writing MOC
    print("\n🔗 Syncing with Featured Writing MOC...")
    generate_featured_moc_file(VAULT_ROOT, generated_episodes)
    
    print("\n✨ Writing sync pipeline finished perfectly!")

if __name__ == "__main__":
    sync_writings()
